import openpyxl

def read_excel():
    excel = openpyxl.load_workbook("data_qc_配置信息.xlsx")

    sheet = excel['data_qc_recordsnew']

    rows = sheet.max_row
    cols = sheet.max_column

    for r in range(1,rows+1):
        for c in range(1,cols+1):
            item = sheet.cell(r,c).value
            print(" %s\t"%item,end='\n')

if __name__ == '__main__':
    read_excel()

print("read done.")